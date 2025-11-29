# -*- coding: utf-8 -*-
from typing import Union
from pathlib import Path
import pandas as pd

def build_observed_q(
    input_path,                 # 源文件(.xlsx/.xls/.csv)
    output_path,                # 目标文件(.xlsx 或 .csv)
    station_id,                 # 站号(如 322)
    value_col=None,             # 数值列名，不传则取最后一列
    sheet_name=0                # Excel 的工作表名或索引
):
    in_path = Path(input_path)

    # 读取源数据（优先使用 openpyxl 读 xlsx）
    if in_path.suffix.lower() in (".xlsx", ".xls"):
        try:
            df = pd.read_excel(str(in_path), sheet_name=sheet_name, engine="openpyxl")
        except Exception:
            # .xls 时可回退为默认引擎（需要 xlrd）
            df = pd.read_excel(str(in_path), sheet_name=sheet_name)
    else:
        df = pd.read_csv(str(in_path))

    # 统一列名
    df.columns = [str(c).strip() for c in df.columns]

    # 识别 年 / 月 / 日 列（中英文都尝试）
    def find_col(keys):
        for c in df.columns:
            sc = str(c).lower()
            for k in keys:
                if k in sc:
                    return c
        return None

    year_col  = find_col(["年", "year"])
    month_col = find_col(["月", "month"])
    day_col   = find_col(["日", "day"])

    if not (year_col and month_col and day_col):
        cols = list(df.columns)
        if len(cols) >= 3:
            year_col, month_col, day_col = cols[:3]
        else:
            raise ValueError("无法识别 年/月/日 列，请检查源表头。")

    # 值列：优先使用传入的 value_col，否则取最后一列（排除年月日）
    if (value_col is None) or (value_col not in df.columns):
        candidates = [c for c in df.columns if c not in (year_col, month_col, day_col)]
        if not candidates:
            raise ValueError("未找到数值列。")
        value_col = candidates[-1]

    # 仅保留有效数字行
    def is_num(x):
        try:
            float(x)
            return True
        except Exception:
            return False

    df = df[
        df[year_col].apply(is_num) &
        df[month_col].apply(is_num) &
        df[day_col].apply(is_num) &
        df[value_col].apply(is_num)
    ].copy()

    # 组装日期
    df["year"]  = df[year_col].astype(int)
    df["month"] = df[month_col].astype(int)
    df["day"]   = df[day_col].astype(int)
    dt = pd.to_datetime(df[["year", "month", "day"]], errors="coerce")
    df = df[~dt.isna()].copy()
    df["DATETIME"] = dt

    # 格式化为 "YYYY/M/D H:MM:SS"（不补零的月/日，小时不补零，分秒两位）
    def fmt_ts(ts):
        return "{}/{}/{} {}:{:02d}:{:02d}".format(
            ts.year, ts.month, ts.day, ts.hour, ts.minute, ts.second
        )

    df["DT_STR"] = df["DATETIME"].apply(fmt_ts)

    # 组装输出行
    rows = list(zip(
        [station_id] * len(df),
        df["DT_STR"],
        ["Q"] * len(df),
        df[value_col].astype(float)
    ))

    out_path = Path(output_path)
    if out_path.suffix.lower() in (".xlsx", ".xls"):
        # 用 openpyxl 手写 Excel，保证首行 "#UTCTIME" 与表头格式
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"

        ws.cell(row=1, column=1, value="#UTCTIME")
        headers = ["StationID", "DATETIME", "Type", "VALUE"]
        for j, h in enumerate(headers, start=1):
            ws.cell(row=2, column=j, value=h)

        for i, (sid, dtstr, typ, val) in enumerate(rows, start=3):
            ws.cell(row=i, column=1, value=sid)
            ws.cell(row=i, column=2, value=dtstr)  # 作为字符串写入，显示为 2008/1/1 0:00:00
            ws.cell(row=i, column=3, value=typ)
            ws.cell(row=i, column=4, value=float(val))

        wb.save(str(out_path))
    else:
        # 写 CSV：第一行 "#UTCTIME"，第二行表头
        with open(str(out_path), "w", encoding="utf-8") as f:
            f.write("#UTCTIME\n")
            f.write("StationID,DATETIME,Type,VALUE\n")
            for sid, dtstr, typ, val in rows:
                f.write("{},{},{},{}\n".format(sid, dtstr, typ, val))

    # 返回一个 DataFrame 预览（不含第一行 "#UTCTIME"）
    return pd.DataFrame(rows, columns=["StationID", "DATETIME", "Type", "VALUE"])

if __name__ == '__main__':
    # 针对国家地球系统数据中心下载的水文年鉴格式
    input_dir = r"G:\program\seims\SEIMS_HAND\data\poyang_lake1\data_prepare\observed\外洲站径流数据\鄱阳湖流域外洲站日流量数据集（2013-2019年）\鄱阳湖流量.xlsx"
    out_csv   = r"G:\program\seims\SEIMS_HAND\data\poyang_lake1\data_prepare\observed\observed_Q_322.csv"
    build_observed_q(
        input_path=input_dir,   # 图1格式的源表
        output_path=out_csv,
        station_id=322,                     # 你指定的站号
        value_col="外洲"                    # 若不写，则默认用最后一列
    )
